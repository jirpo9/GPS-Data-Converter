#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
GPS Data Converter

Tento skript převádí GPS data z CSV formátu do formátovaného XLSX souboru.
Vstupem je CSV soubor s GPS daty, výstupem je XLSX soubor s několika listy
obsahujícími různé pohledy na data: původní data, převedené časové značky,
rychlost, sprinty a souhrny sprintů.
"""

import os
import sys
import pandas as pd
from openpyxl import load_workbook
import argparse
from datetime import datetime

def convert_csv_to_xlsx(input_file, output_file=None):
    """
    Převede CSV soubor s GPS daty do formátovaného XLSX souboru.
    
    Args:
        input_file (str): Cesta k vstupnímu CSV souboru
        output_file (str, optional): Cesta k výstupnímu XLSX souboru. Pokud není zadána,
                                     použije se název vstupního souboru s příponou .xlsx
    
    Returns:
        str: Cesta k výstupnímu XLSX souboru
    """
    print(f"Zpracovávám soubor: {input_file}")
    
    # Určení názvu výstupního souboru, pokud není zadán
    if output_file is None:
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        output_file = f"{base_name}_upraveno_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    try:
        # Načtení CSV souboru
        data = pd.read_csv(input_file)
        
        # Odstranění mezer z názvů sloupců
        data.columns = data.columns.str.strip()
        
        print(f"Načteno {len(data)} řádků dat.")
        print(f"Detekované sloupce: {', '.join(data.columns)}")
        
        # Vytvoření nového Excel souboru s formátováním
        detailed_sprints = []  # Inicializace prázdného seznamu pro sběr dat
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # List 1: Původní data - zachování původního formátu
            data.to_excel(writer, sheet_name='Puvodni data', index=False)
            print("Vytvořen list: Puvodni data")
            
            # List 2: Úprava 1 - Převod data a času do čitelného formátu a rychlosti na km/h
            uprava1 = data.copy()
            if 'Excel Timestamp' in uprava1.columns:
                uprava1['Excel Timestamp'] = pd.to_datetime(uprava1['Excel Timestamp'], origin='1899-12-30', unit='d')
            else:
                print("Varování: Sloupec 'Excel Timestamp' nebyl nalezen. Zkontrolujte vstupní data.")
            
            if 'Speed' in uprava1.columns:
                uprava1['Speed'] = uprava1['Speed'] * 3.6  # Převod rychlosti z m/s na km/h
                uprava1['Speed'] = uprava1['Speed'].map(lambda x: f'{x:.3f}')  # Formátování rychlosti na tři desetinná místa
            else:
                print("Varování: Sloupec 'Speed' nebyl nalezen. Zkontrolujte vstupní data.")
            
            # Vybereme pouze sloupce 'Excel Timestamp' a 'Speed', pokud existují
            existing_columns = [col for col in ['Excel Timestamp', 'Speed'] if col in uprava1.columns]
            uprava1 = uprava1[existing_columns]
            
            # Formátování Excel Timestamp pro zobrazení sekund se třemi desetinnými místy
            if 'Excel Timestamp' in uprava1.columns:
                uprava1['Excel Timestamp'] = uprava1['Excel Timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S.%f').str[:-3]
            
            uprava1.to_excel(writer, sheet_name='Cas a Rychlost', index=False)
            print("Vytvořen list: Cas a Rychlost")
            
            # List 3: Filtrovaná úprava - filtr pro rychlost >= 25,2 km/h a výpočet segmentů (sprintů)
            if 'Speed' in uprava1.columns and 'Excel Timestamp' in uprava1.columns:
                uprava3 = uprava1[['Excel Timestamp', 'Speed']].dropna().copy()
                uprava3 = uprava3[uprava3['Speed'].astype(float) >= 25.2]
                
                if not uprava3.empty:
                    # Přidání sloupce segmentu na základě kontinuální rychlosti sprintu
                    uprava3['Excel Timestamp'] = pd.to_datetime(uprava3['Excel Timestamp'], errors='coerce')
                    uprava3['Sprint'] = (uprava3['Excel Timestamp'].diff().dt.total_seconds() > 5).cumsum() + 1
                    
                    # Výpočet doby trvání pro každý řádek
                    for sprint, group in uprava3.groupby('Sprint'):
                        # Získání počátečního a koncového času pro aktuální sprint
                        start_time = group['Excel Timestamp'].min()
                        end_time = group['Excel Timestamp'].max()
                        
                        # Celková doba trvání pro aktuální sprint
                        total_duration = (end_time - start_time).total_seconds()
                        
                        for i in range(len(group)):
                            if i == len(group) - 1:  # Poslední záznam, nastavit zbývající trvání na 0
                                time_duration = 0
                            else:  # Pro všechny ostatní záznamy vypočítat čas do dalšího záznamu
                                time_duration = (group['Excel Timestamp'].iloc[i + 1] - group['Excel Timestamp'].iloc[i]).total_seconds()
                            
                            # Přidání aktuálních dat řádku do seznamu detailed_sprints
                            detailed_sprints.append({
                                'Excel Timestamp': group['Excel Timestamp'].iloc[i],
                                'Speed': group['Speed'].iloc[i],
                                'Sprint': sprint,  # Použití čísla sprintu pro řazení
                                'Cas_trvani_sprintu_s': f'{time_duration:.3f}'  # Formátování času na 3 desetinná místa
                            })
                    
                    detailed_sprints_df = pd.DataFrame(detailed_sprints)  # Vytvoření DataFrame ze seznamu
                    detailed_sprints_df['Excel Timestamp'] = pd.to_datetime(detailed_sprints_df['Excel Timestamp'], errors='coerce')
                    detailed_sprints_df['Speed'] = detailed_sprints_df['Speed'].astype(float).map(lambda x: f'{x:.3f}')
                    
                    # Formátování Excel Timestamp pro zobrazení sekund se třemi desetinnými místy
                    detailed_sprints_df['Excel Timestamp'] = detailed_sprints_df['Excel Timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S.%f').str[:-3]
                    
                    # Seřazení podle čísla sprintu a času Excel Timestamp
                    detailed_sprints_df = detailed_sprints_df.sort_values(by=['Sprint', 'Excel Timestamp']).reset_index(drop=True)
                    detailed_sprints_df['Sprint'] = 'Sprint ' + detailed_sprints_df['Sprint'].astype(str)
                    
                    detailed_sprints_df.to_excel(writer, sheet_name='>=25,2_kmh', index=False)
                    print("Vytvořen list: >=25,2_kmh")
                    
                    # List 4: Souhrn sprintů bez rychlosti
                    segments = uprava3.groupby('Sprint').agg(
                        Start_time=('Excel Timestamp', 'min'),
                        End_time=('Excel Timestamp', 'max'),
                        Total_time=('Excel Timestamp', lambda x: (x.max() - x.min()).total_seconds())
                    ).reset_index()
                    
                    segments['Start_time'] = segments['Start_time'].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3] if pd.notnull(x) and x is not pd.NaT else '')
                    segments['End_time'] = segments['End_time'].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3] if pd.notnull(x) and x is not pd.NaT else '')
                    segments = segments[['Sprint', 'Start_time', 'End_time', 'Total_time']]
                    
                    # Formátování Total_time na tři desetinná místa jako řetězec pro konzistentní výstup
                    segments['Total_time'] = segments['Total_time'].astype(float).map(lambda x: f'{x:.3f}')
                    
                    segments['Sprint'] = segments['Sprint'].astype(int)  # Převod zpět na celé číslo pro řazení
                    segments = segments.sort_values('End_time').reset_index(drop=True)
                    segments['Sprint'] = 'Sprint ' + segments['Sprint'].astype(str)
                    segments.to_excel(writer, sheet_name='Pocet sprintu', index=False)
                    print("Vytvořen list: Pocet sprintu")
                    
                    # List 5: Sprinty s dobou trvání >= 1s
                    sprint_1s = segments[segments['Total_time'].astype(float) >= 1.0]
                    sprint_1s.to_excel(writer, sheet_name='>=1s', index=False)
                    print("Vytvořen list: >=1s")
                else:
                    print("Varování: Žádná data nesplňují kritérium rychlosti >= 25,2 km/h.")
            else:
                print("Varování: Některé potřebné sloupce chybí. Nelze vytvořit listy pro sprinty.")
        
        # Úprava šířky sloupců tak, aby odpovídaly textu ve všech buňkách a listech
        wb = load_workbook(output_file)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for column_cells in sheet.columns:
                max_length = 0
                for cell in column_cells:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
        
        wb.save(output_file)
        print(f"Soubor byl úspěšně uložen: {output_file}")
        
        return output_file
    
    except Exception as e:
        print(f"Chyba při zpracování souboru: {e}")
        return None

def main():
    """
    Hlavní funkce programu - zpracuje argumenty příkazové řádky a spustí konverzi.
    """
    parser = argparse.ArgumentParser(description='Konverze GPS dat z CSV do formátovaného XLSX.')
    parser.add_argument('input_file', help='Cesta k vstupnímu CSV souboru')
    parser.add_argument('-o', '--output', help='Cesta k výstupnímu XLSX souboru (volitelné)')
    
    args = parser.parse_args()
    
    # Kontrola existence vstupního souboru
    if not os.path.isfile(args.input_file):
        print(f"Chyba: Vstupní soubor '{args.input_file}' nebyl nalezen.")
        return 1
    
    # Volání funkce pro konverzi
    output_file = convert_csv_to_xlsx(args.input_file, args.output)
    
    if output_file:
        return 0
    else:
        return 1

if __name__ == "__main__":
    sys.exit(main())